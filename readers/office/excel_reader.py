"""
Excel 文档读取器
"""
from pathlib import Path

from core.file_reader import BaseFileReader, FileContent
from config import CHUNK_SIZE, CHUNK_OVERLAP


class ExcelReader(BaseFileReader):
    """Excel 读取器 (.xlsx, .xls)"""
    
    def __init__(self):
        super().__init__()
        self.supported_extensions = [".xlsx", ".xls"]
    
    def read(self, file_path: Path) -> FileContent:
        """读取 Excel 文档"""
        file_path = Path(file_path)
        
        wb = None
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, data_only=True)
            
            all_sheets = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_rows = [f"=== 工作表: {sheet_name} ==="]
                
                for row in sheet.iter_rows():
                    row_values = []
                    for cell in row:
                        if cell.value is not None:
                            row_values.append(str(cell.value))
                    
                    if row_values:
                        sheet_rows.append(" | ".join(row_values))
                
                all_sheets.append("\n".join(sheet_rows))
            
            content = "\n\n".join(all_sheets)
            title = file_path.stem
            
        except Exception as e:
            content = f"[Error reading Excel: {e}]"
            title = file_path.stem
        
        # 提取元数据
        metadata = self.extract_metadata_from_path(file_path)
        metadata["sheet_count"] = len(wb.sheetnames) if wb else 0
        
        # 分块
        chunks = self.chunk_content(content, CHUNK_SIZE, CHUNK_OVERLAP)
        
        return FileContent(
            source=str(file_path.absolute()),
            content=content,
            file_type="excel",
            title=title,
            metadata=metadata,
            chunks=chunks,
            file_hash=self.calculate_hash(content)
        )
